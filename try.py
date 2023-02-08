import os
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import urllib.request
from html_table_parser.parser import HTMLTableParser
from googletrans import Translator
from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl.utils.dataframe import dataframe_to_rows

def appendfunc(df,key):
    path=r"C:\Users\Ayush Gupta\OneDrive\Desktop\IMARC\Pandas\Scraping\100 PPI MAIN SHEET (ALL DATES).xlsx"
    n=len(key)
    key=key[:31]

    if os.path.isfile(path):
        workbook=openpyxl.load_workbook(path)
        
        if not key in workbook.sheetnames:
            workbook.create_sheet(key)
            sheet=workbook[key]
            for row in dataframe_to_rows(df,header=True,index=False):
                sheet.append(row)
            workbook.save(path)
        
        else:            
            sheet=workbook[key]
            for row in dataframe_to_rows(df,header=False,index=False):
                sheet.append(row)
            workbook.save(path)
        workbook.close()       

    else:
        with pd.ExcelWriter(path,engine='openpyxl') as writer:
            df.to_excel(writer,index=False,sheet_name=key)

def url_get_contents(url):
    req = urllib.request.Request(url=url)
    f = urllib.request.urlopen(req)
    return f.read()

xhtml=url_get_contents('https://futures.100ppi.com').decode('utf-8')
p = HTMLTableParser()
p.feed(xhtml)

urlpage=urlopen("https://futures.100ppi.com").read()
bswebpage=BeautifulSoup(urlpage)

translator=Translator()

#Table title
results=bswebpage.find_all("div",{'class':"fr width220 height_28p height28p t12 greyk textright paddr10"})

#Table date
results2=bswebpage.find_all("div",{'class':"fl width250 futtb1 t14"})

dfa=pd.DataFrame(p.tables[3])
resulta=results[0]
resulta=translator.translate(resulta.contents[0])
resulta2=results2[0]
dfa=dfa.applymap(lambda value: translator.translate(value).text)
dfa['Date'] = resulta.text[13:23]
dfa['Currency'] = resulta.text[-3:]
resulta2=translator.translate(resulta2.text)
appendfunc(dfa,resulta2.text)

dfb=pd.DataFrame(p.tables[4])
resultb=results[1]
resultb=translator.translate(resultb.contents[0])
resultb2=results2[1]
dfb=dfb.applymap(lambda value: translator.translate(value).text)
dfb['Date'] = resultb.text[13:23]
dfb['Currency'] = resultb.text[-3:]
resultb2=translator.translate(resultb2.text)
appendfunc(dfb,resultb2.text)

dfc=pd.DataFrame(p.tables[5])
resultc=results[2]
resultc=translator.translate(resultc.contents[0])
resultc2=results2[2]
dfc=dfc.applymap(lambda value: translator.translate(value).text)
dfc['Date'] = resultc.text[13:23]
dfc['Currency'] = resultc.text[-3:]
resultc2=translator.translate(resultc2.text)
appendfunc(dfc,resultc2.text)

dfd=pd.DataFrame(p.tables[6])
resultd=results[3]
resultd=translator.translate(resultd.contents[0])
resultd2=results2[3]
dfd=dfd.applymap(lambda value: translator.translate(value).text)
dfd['Date'] = resultd.text[13:23]
dfd['Currency'] = resultd.text[-3:]
resultd2=translator.translate(resultd2.text)
appendfunc(dfd,resultd2.text)

#Remove duplicates
path=r"C:\Users\Ayush Gupta\OneDrive\Desktop\IMARC\Pandas\Scraping\100 PPI MAIN SHEET (ALL DATES).xlsx"
dfnew=pd.read_excel(path,sheet_name=None)
for name,sheet in dfnew.items():
    df1=pd.read_excel(path,sheet_name=name)
    df1=df1.drop_duplicates(subset=[0,6,'Date'],keep="first")

    with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer: 
        workBook = writer.book
        try:
            workBook.remove(workBook[name])
        finally:
            df1.to_excel(writer, sheet_name=name,index=False)
            writer.save()
